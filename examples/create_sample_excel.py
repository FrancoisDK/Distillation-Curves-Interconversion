from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create D86 sample data
d86_data = {
    'Volume %': [5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95],
    'Temperature C': [150, 175, 210, 245, 280, 315, 350, 385, 420, 455, 475],
    'Density kg/m3': [850, 850, 850, 850, 850, 850, 850, 850, 850, 850, 850]
}

# Create Excel workbook manually for better control
wb = Workbook()

# D86 sheet
ws = wb.active
ws.title = "D86"

headers = ['Volume %', 'Temperature C', 'Density kg/m3']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data
for row_idx, (vol, temp, dens) in enumerate(zip(d86_data['Volume %'], 
                                                   d86_data['Temperature C'],
                                                   d86_data['Density kg/m3']), 2):
    ws.cell(row=row_idx, column=1).value = vol
    ws.cell(row=row_idx, column=2).value = temp
    ws.cell(row=row_idx, column=3).value = dens

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15

# D2887 sheet
ws2 = wb.create_sheet("D2887")

d2887_data = {
    'Volume %': [5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95],
    'Temperature C': [135, 160, 195, 230, 265, 300, 335, 370, 405, 440, 460],
    'Density kg/m3': [880, 880, 880, 880, 880, 880, 880, 880, 880, 880, 880]
}

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data
for row_idx, (vol, temp, dens) in enumerate(zip(d2887_data['Volume %'], 
                                                   d2887_data['Temperature C'],
                                                   d2887_data['Density kg/m3']), 2):
    ws2.cell(row=row_idx, column=1).value = vol
    ws2.cell(row=row_idx, column=2).value = temp
    ws2.cell(row=row_idx, column=3).value = dens

# Adjust column widths
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15

# Save file
wb.save('sample_distillation_data.xlsx')
print("✅ Created sample_distillation_data.xlsx with D86 and D2887 sheets")
