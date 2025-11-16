"""
Distillation Curve Interconversion GUI
Converts between ASTM D86, ASTM D2887, and TBP distillation curves
"""

import sys
from pathlib import Path
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                                QGroupBox, QLabel, QComboBox, QPushButton, QTableWidget, 
                                QTableWidgetItem, QHeaderView, QMessageBox, QFileDialog,
                                QCheckBox, QDoubleSpinBox, QFormLayout, QSpacerItem, 
                                QSizePolicy, QTabWidget, QMenu)
from PySide6.QtCore import Qt, QEvent, QPoint
from PySide6.QtGui import QFont, QKeyEvent, QAction
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from matplotlib.ticker import AutoMinorLocator, MaxNLocator
import matplotlib.pyplot as plt
import numpy as np

# Import the Oil class from bp_conversions
import sys
sys.path.append(str(Path(__file__).parent))
from bp_conversions import Oil


class InteractiveTableWidget(QTableWidget):
    """Custom table widget with Enter key navigation and Excel paste support"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # Enable context menu
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
    
    def show_context_menu(self, position: QPoint):
        """Show right-click context menu with Copy and Paste options"""
        menu = QMenu(self)
        
        # Copy action
        copy_action = QAction("üìã Copy", self)
        copy_action.setShortcut("Ctrl+C")
        copy_action.triggered.connect(self.copy_to_clipboard)
        menu.addAction(copy_action)
        
        # Paste action
        paste_action = QAction("üìÑ Paste", self)
        paste_action.setShortcut("Ctrl+V")
        paste_action.triggered.connect(self.paste_from_clipboard)
        menu.addAction(paste_action)
        
        menu.addSeparator()
        
        # Clear selection action
        clear_action = QAction("üóëÔ∏è Clear", self)
        clear_action.triggered.connect(self.clear_selection)
        menu.addAction(clear_action)
        
        # Show menu at cursor position
        menu.exec(self.viewport().mapToGlobal(position))
    
    def copy_to_clipboard(self):
        """Copy selected cells to clipboard in Excel format"""
        selection = self.selectedRanges()
        
        if not selection:
            return
        
        # Get the first selection range
        sel_range = selection[0]
        
        # Build tab-separated text
        rows = []
        for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
            row_data = []
            for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                item = self.item(row, col)
                row_data.append(item.text() if item else "")
            rows.append("\t".join(row_data))
        
        # Join rows with newlines
        text = "\n".join(rows)
        
        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
    
    def clear_selection(self):
        """Clear the content of selected cells"""
        selection = self.selectedRanges()
        
        if not selection:
            return
        
        sel_range = selection[0]
        
        for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
            for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                item = self.item(row, col)
                if item:
                    item.setText("")
        
    def keyPressEvent(self, event: QKeyEvent):
        """Handle key press events for navigation and paste"""
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            # Move to next row when Enter is pressed
            current_row = self.currentRow()
            current_col = self.currentColumn()
            
            if current_row < self.rowCount() - 1:
                # Move to next row, same column
                self.setCurrentCell(current_row + 1, current_col)
            else:
                # If on last row, add a new row and move to it
                self.insertRow(self.rowCount())
                self.setCurrentCell(current_row + 1, current_col)
                
                # Add empty items to the new row
                for col in range(self.columnCount()):
                    if not self.item(current_row + 1, col):
                        self.setItem(current_row + 1, col, QTableWidgetItem(""))
            
            event.accept()
        elif event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier:
            # Handle Ctrl+V paste from Excel
            self.paste_from_clipboard()
            event.accept()
        elif event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            # Handle Ctrl+C copy to clipboard
            self.copy_to_clipboard()
            event.accept()
        else:
            super().keyPressEvent(event)
    
    def paste_from_clipboard(self):
        """Paste data from clipboard (Excel format)"""
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        
        if not text:
            return
        
        # Split by newlines to get rows
        rows = text.strip().split('\n')
        
        if not rows:
            return
        
        # Get current cell position
        current_row = self.currentRow()
        current_col = self.currentColumn()
        
        if current_row < 0:
            current_row = 0
        if current_col < 0:
            current_col = 0
        
        # Determine how many rows we need
        needed_rows = current_row + len(rows)
        
        # Add rows if necessary
        while self.rowCount() < needed_rows:
            self.insertRow(self.rowCount())
        
        # Parse and insert data
        for row_idx, row_data in enumerate(rows):
            # Split by tab (Excel uses tabs between columns)
            cells = row_data.split('\t')
            
            target_row = current_row + row_idx
            
            for col_idx, cell_value in enumerate(cells):
                target_col = current_col + col_idx
                
                # Only paste within table column bounds
                if target_col < self.columnCount():
                    cell_value = cell_value.strip()
                    
                    # Create item if it doesn't exist
                    if not self.item(target_row, target_col):
                        self.setItem(target_row, target_col, QTableWidgetItem(cell_value))
                    else:
                        self.item(target_row, target_col).setText(cell_value)


class DistillationConverterGUI(QMainWindow):
    """Main GUI window for distillation curve interconversion"""
    
    # Standard distillation points
    STANDARD_POINTS = [0, 10, 30, 50, 70, 90, 95, 100]
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Distillation Curve Interconversion - D86 / D2887 / TBP")
        self.setGeometry(100, 100, 1400, 900)
        
        # Data storage
        self.input_data = {}
        self.oil_object = None
        self.density = 800.0  # Default density kg/m¬≥
        
        # Create menu bar
        self.create_menu_bar()
        
        # Create main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        
        # Left panel for input
        left_panel = self.create_left_panel()
        main_layout.addWidget(left_panel, stretch=2)
        
        # Right panel for output
        right_panel = self.create_right_panel()
        main_layout.addWidget(right_panel, stretch=3)
        
        # Initialize table with standard points
        self.update_input_table()
    
    def create_menu_bar(self):
        """Create menu bar with Help section"""
        menubar = self.menuBar()
        
        # Help menu
        help_menu = menubar.addMenu("‚ùì &Help")
        
        # About action
        about_action = help_menu.addAction("‚ÑπÔ∏è About")
        about_action.triggered.connect(self.show_about)
        
        # Methods documentation
        methods_action = help_menu.addAction("üî¨ Conversion Methods")
        methods_action.triggered.connect(self.show_methods_help)
        
        # API correlations
        api_action = help_menu.addAction("üìä API Correlations")
        api_action.triggered.connect(self.show_api_help)
    
    def show_about(self):
        """Show about dialog"""
        about_text = """
        <h2>üß™ Distillation Curve Interconversion Tool</h2>
        <p><b>Version:</b> 1.0</p>
        <p><b>Purpose:</b> Convert between different petroleum distillation curve types:</p>
        <ul>
            <li>üîµ ASTM D86 (Atmospheric Distillation)</li>
            <li>üî¥ ASTM D2887 (Simulated Distillation by GC)</li>
            <li>üü¢ TBP (True Boiling Point)</li>
        </ul>
        <p><b>Methods Used:</b></p>
        <ul>
            <li>üìä API Technical Data Book (1997) correlations</li>
            <li> PCHIP (Piecewise Cubic Hermite Interpolating Polynomial) for smooth curves</li>
        </ul>
        <p><b>Developer:</b> ‚öôÔ∏è Petroleum Engineering Application</p>
        """
        QMessageBox.about(self, "About", about_text)
    
    def show_methods_help(self):
        """Show conversion methods help"""
        methods_text = """
        <h2>Conversion Methods Overview</h2>
        
        <h3>1. ASTM D86 (Atmospheric Distillation)</h3>
        <p>Physical distillation method performed at atmospheric pressure. 
        Subject to apparatus effects, reflux, and operational variations.</p>
        <ul>
            <li><b>Basis:</b> Volume % distilled</li>
            <li><b>Typical Use:</b> Gasoline, diesel, jet fuel specifications</li>
        </ul>
        
        <h3>2. ASTM D2887 (Simulated Distillation)</h3>
        <p>Gas chromatography-based method that simulates true boiling point distillation.
        More repeatable than D86, provides detailed boiling range distribution.</p>
        <ul>
            <li><b>Basis:</b> Weight % distilled</li>
            <li><b>Typical Use:</b> Refinery process simulation, crude assays</li>
            <li><b>Relationship:</b> D2887 ‚âà TBP (very similar)</li>
        </ul>
        
        <h3>3. TBP (True Boiling Point)</h3>
        <p>Theoretical distillation with infinite reflux and infinite plates.
        Represents true component separation by boiling point.</p>
        <ul>
            <li><b>Basis:</b> Volume % distilled</li>
            <li><b>Typical Use:</b> Process design, simulation models</li>
            <li><b>Method:</b> API Technical Data Book (1997) correlations</li>
        </ul>
        
        <h3>Interpolation Method</h3>
        <p><b>PCHIP (Piecewise Cubic Hermite Interpolating Polynomial)</b></p>
        <ul>
            <li>Preserves monotonicity (temperatures always increase)</li>
            <li>No overshoots or oscillations between data points</li>
            <li>Shape-preserving with C¬π continuity</li>
            <li>Ideal for physical data like distillation curves</li>
        </ul>
        """
        msg = QMessageBox(self)
        msg.setWindowTitle("Conversion Methods")
        msg.setTextFormat(Qt.RichText)
        msg.setText(methods_text)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()
    
    def show_api_help(self):
        """Show API method help"""
        api_text = """
        <h2>API Technical Data Book Methods</h2>
        
        <h3>API D86 to TBP Conversion</h3>
        <p>Uses power-law correlations: <b>TBP = a √ó D86<sup>b</sup></b></p>
        <p>where a and b are constants that vary with volume % distilled.</p>
        
        <table border="1" cellpadding="5" style="border-collapse: collapse;">
        <tr><th>Vol %</th><th>a</th><th>b</th></tr>
        <tr><td>0 (IBP)</td><td>0.9167</td><td>1.0019</td></tr>
        <tr><td>10</td><td>0.5277</td><td>1.0900</td></tr>
        <tr><td>30</td><td>0.7429</td><td>1.0425</td></tr>
        <tr><td>50</td><td>0.8920</td><td>1.0176</td></tr>
        <tr><td>70</td><td>0.8705</td><td>1.0226</td></tr>
        <tr><td>90</td><td>0.9490</td><td>1.0110</td></tr>
        <tr><td>95</td><td>0.8008</td><td>1.0355</td></tr>
        </table>
        
        <h3>API Procedure 3A3.2 (SimDist to D86)</h3>
        <p><b>Step 1:</b> Calculate ASTM(50) from SD(50)</p>
        <p style="margin-left: 20px;"><i>ASTM(50) = 0.77601 √ó SD(50)<sup>1.0395</sup></i></p>
        
        <p><b>Step 2:</b> Calculate temperature differences U<sub>i</sub></p>
        <p style="margin-left: 20px;"><i>U<sub>i</sub> = C √ó T<sub>i</sub><sup>D</sup></i></p>
        <p style="margin-left: 20px;">where T<sub>i</sub> is the temperature difference between cut points</p>
        
        <p><b>Step 3:</b> Build D86 curve from 50% point:</p>
        <ul style="margin-left: 20px;">
            <li>ASTM(0) = ASTM(50) - U‚ÇÑ - U‚ÇÖ - U‚ÇÜ</li>
            <li>ASTM(10) = ASTM(50) - U‚ÇÑ - U‚ÇÖ</li>
            <li>ASTM(30) = ASTM(50) - U‚ÇÑ</li>
            <li>ASTM(70) = ASTM(50) + U‚ÇÉ</li>
            <li>ASTM(90) = ASTM(50) + U‚ÇÉ + U‚ÇÇ</li>
            <li>ASTM(100) = ASTM(50) + U‚ÇÉ + U‚ÇÇ + U‚ÇÅ</li>
        </ul>
        
        <h3>D86 to D2887 Conversion</h3>
        <p>This tool implements the <b>inverse of Procedure 3A3.2:</b></p>
        <ol>
            <li>Calculate SD(50) from D86(50) using inverse equation</li>
            <li>Use weighted interpolation between TBP and D86 for other points</li>
            <li>SimDist positioned ~85% toward TBP from D86</li>
        </ol>
        
        <p><b>Reference:</b> API Technical Data Book - Petroleum Refining, 6th Edition (1997)</p>
        """
        msg = QMessageBox(self)
        msg.setWindowTitle("API Correlations")
        msg.setTextFormat(Qt.RichText)
        msg.setText(api_text)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()
    
    def create_left_panel(self):
        """Create the left panel with input controls"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        
        # Title
        title = QLabel("‚öôÔ∏è Input Configuration")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Input type selection
        input_group = QGroupBox("üìù Select Input Type")
        input_layout = QVBoxLayout()
        
        self.input_type_combo = QComboBox()
        self.input_type_combo.addItems(["üîµ D86 (Atmospheric Distillation)", 
                                        "üî¥ D2887 (SimDis - GC)", 
                                        "üü¢ TBP (True Boiling Point)"])
        self.input_type_combo.currentIndexChanged.connect(self.on_input_type_changed)
        input_layout.addWidget(QLabel("Input Distillation Type:"))
        input_layout.addWidget(self.input_type_combo)
        
        # Basis selection (vol% or wt%)
        self.basis_combo = QComboBox()
        self.basis_combo.addItems(["üìä Volume %", "‚öñÔ∏è Weight %"])
        self.basis_combo.currentIndexChanged.connect(self.on_basis_changed)
        input_layout.addWidget(QLabel("Input Basis:"))
        input_layout.addWidget(self.basis_combo)
        
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)
        
        # Density input
        density_group = QGroupBox("üß™ Material Properties")
        density_layout = QFormLayout()
        
        self.density_spinbox = QDoubleSpinBox()
        self.density_spinbox.setRange(600.0, 1200.0)
        self.density_spinbox.setValue(800.0)
        self.density_spinbox.setSuffix(" kg/m¬≥")
        self.density_spinbox.setDecimals(1)
        self.density_spinbox.valueChanged.connect(self.on_density_changed)
        density_layout.addRow("Density:", self.density_spinbox)
        
        density_group.setLayout(density_layout)
        layout.addWidget(density_group)
        
        # Data input table
        table_group = QGroupBox("üìã Distillation Data Input")
        table_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel("Enter temperature values in ¬∞C for each distillation point:")
        instructions.setWordWrap(True)
        table_layout.addWidget(instructions)
        
        # Instructions for paste
        paste_instructions = QLabel("üí° Tip: Press Enter to move to next row. Right-click for Copy/Paste menu. Ctrl+C to copy, Ctrl+V to paste.")
        paste_instructions.setWordWrap(True)
        paste_instructions.setStyleSheet("font-style: italic; color: gray;")
        table_layout.addWidget(paste_instructions)
        
        # Table widget - using custom interactive table
        self.input_table = InteractiveTableWidget()
        self.input_table.setColumnCount(2)
        self.input_table.setHorizontalHeaderLabels(["Vol %", "Temperature (¬∞C)"])
        self.input_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.input_table.verticalHeader().setVisible(False)
        self.input_table.cellChanged.connect(self.on_cell_changed)
        self.input_table.setSelectionMode(QTableWidget.ExtendedSelection)  # Allow multi-cell selection for copy/paste
        
        table_layout.addWidget(self.input_table)
        
        # Buttons for table management - use native styling
        btn_layout = QHBoxLayout()
        
        self.add_point_btn = QPushButton("‚ûï Add Point")
        self.add_point_btn.clicked.connect(self.add_data_point)
        btn_layout.addWidget(self.add_point_btn)
        
        self.remove_point_btn = QPushButton("‚ûñ Remove Point")
        self.remove_point_btn.clicked.connect(self.remove_data_point)
        btn_layout.addWidget(self.remove_point_btn)
        
        self.clear_btn = QPushButton("üóëÔ∏è Clear All")
        self.clear_btn.clicked.connect(self.clear_all_data)
        btn_layout.addWidget(self.clear_btn)
        
        table_layout.addLayout(btn_layout)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group, stretch=1)
        
        # Conversion controls
        convert_group = QGroupBox("üîÑ Conversion Options")
        convert_layout = QVBoxLayout()
        
        self.convert_d86_check = QCheckBox("üîµ Convert to D86")
        self.convert_d86_check.setChecked(True)
        convert_layout.addWidget(self.convert_d86_check)
        
        self.convert_d2887_check = QCheckBox("üî¥ Convert to D2887 (SimDis)")
        self.convert_d2887_check.setChecked(True)
        convert_layout.addWidget(self.convert_d2887_check)
        
        self.convert_tbp_check = QCheckBox("üü¢ Convert to TBP")
        self.convert_tbp_check.setChecked(True)
        convert_layout.addWidget(self.convert_tbp_check)
        
        convert_group.setLayout(convert_layout)
        layout.addWidget(convert_group)
        
        # Calculate button - use native styling with minimal customization
        self.calculate_btn = QPushButton("üßÆ Calculate Conversions")
        self.calculate_btn.clicked.connect(self.calculate_conversions)
        self.calculate_btn.setMinimumHeight(40)
        self.calculate_btn.setDefault(True)  # Make it the default button
        layout.addWidget(self.calculate_btn)
        
        # Import buttons - use native styling
        import_layout = QHBoxLayout()
        
        self.import_csv_btn = QPushButton("üì• Import CSV")
        self.import_csv_btn.clicked.connect(self.import_csv)
        import_layout.addWidget(self.import_csv_btn)
        
        self.import_excel_btn = QPushButton("üì• Import Excel")
        self.import_excel_btn.clicked.connect(self.import_excel)
        import_layout.addWidget(self.import_excel_btn)
        
        layout.addLayout(import_layout)
        
        # Export buttons - use native styling
        export_layout = QVBoxLayout()
        
        # First row of export buttons
        export_row1 = QHBoxLayout()
        self.export_csv_btn = QPushButton("üìÑ Export CSV")
        self.export_csv_btn.clicked.connect(self.export_csv)
        self.export_csv_btn.setEnabled(False)
        export_row1.addWidget(self.export_csv_btn)
        
        self.export_excel_btn = QPushButton("üìä Export Excel")
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_excel_btn.setEnabled(False)
        export_row1.addWidget(self.export_excel_btn)
        
        export_layout.addLayout(export_row1)
        
        # Second row - JSON export
        self.export_json_btn = QPushButton("üì¶ Export JSON")
        self.export_json_btn.clicked.connect(self.export_json)
        self.export_json_btn.setEnabled(False)
        export_layout.addWidget(self.export_json_btn)
        
        layout.addLayout(export_layout)
        
        # Add stretch to push everything to top
        layout.addStretch()
        
        return panel
    
    def create_right_panel(self):
        """Create the right panel with output display"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        
        # Title
        title = QLabel("üìà Conversion Results")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Tab widget for different views
        self.tab_widget = QTabWidget()
        
        # Plot tab
        plot_tab = QWidget()
        plot_layout = QVBoxLayout(plot_tab)
        
        # Create matplotlib figure
        self.figure = Figure(figsize=(10, 6))
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, plot_tab)
        
        plot_layout.addWidget(self.toolbar)
        plot_layout.addWidget(self.canvas)
        
        self.tab_widget.addTab(plot_tab, "Plot")
        
        # Results table tab
        table_tab = QWidget()
        table_layout = QVBoxLayout(table_tab)
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(4)
        self.results_table.setHorizontalHeaderLabels(["Vol %", "Input (¬∞C)", "D86 (¬∞C)", "D2887 (¬∞C)"])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.results_table.setSelectionMode(QTableWidget.NoSelection)
        
        table_layout.addWidget(self.results_table)
        
        self.tab_widget.addTab(table_tab, "Data Table")
        
        # Properties tab
        properties_tab = QWidget()
        properties_layout = QVBoxLayout(properties_tab)
        
        self.properties_table = QTableWidget()
        self.properties_table.setColumnCount(2)
        self.properties_table.setHorizontalHeaderLabels(["Property", "Value"])
        self.properties_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.properties_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        properties_layout.addWidget(self.properties_table)
        
        self.tab_widget.addTab(properties_tab, "Properties")
        
        layout.addWidget(self.tab_widget)
        
        return panel
    
    def update_input_table(self):
        """Update the input table with standard points"""
        self.input_table.blockSignals(True)  # Prevent triggering cellChanged
        self.input_table.setRowCount(len(self.STANDARD_POINTS))
        
        for i, vol_pct in enumerate(self.STANDARD_POINTS):
            # Volume/Weight % column
            item = QTableWidgetItem(str(vol_pct))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make read-only
            self.input_table.setItem(i, 0, item)
            
            # Temperature column (editable)
            temp_item = QTableWidgetItem("")
            self.input_table.setItem(i, 1, temp_item)
        
        self.input_table.blockSignals(False)
    
    def on_input_type_changed(self):
        """Handle input type change"""
        input_type = self.input_type_combo.currentText()
        
        # Update basis label based on typical usage
        if "D86" in input_type:
            self.basis_combo.setCurrentIndex(0)  # Volume %
        elif "D2887" in input_type:
            # D2887 can be either, but typically weight %
            pass
        elif "TBP" in input_type:
            # TBP can be either
            pass
    
    def on_basis_changed(self):
        """Handle basis change"""
        basis = self.basis_combo.currentText()
        # Update table header
        if "Volume" in basis:
            self.input_table.setHorizontalHeaderLabels(["Vol %", "Temperature (¬∞C)"])
        else:
            self.input_table.setHorizontalHeaderLabels(["Wt %", "Temperature (¬∞C)"])
    
    def on_density_changed(self, value):
        """Handle density change"""
        self.density = value
    
    def on_cell_changed(self, row, column):
        """Handle cell value change"""
        if column == 1:  # Temperature column
            try:
                vol_pct = float(self.input_table.item(row, 0).text())
                temp_text = self.input_table.item(row, 1).text()
                if temp_text:
                    temp = float(temp_text)
                    self.input_data[vol_pct] = temp
                elif vol_pct in self.input_data:
                    del self.input_data[vol_pct]
            except (ValueError, AttributeError):
                pass
    
    def add_data_point(self):
        """Add a new data point row"""
        row_count = self.input_table.rowCount()
        self.input_table.insertRow(row_count)
        
        # Add editable volume % and temperature cells
        vol_item = QTableWidgetItem("")
        self.input_table.setItem(row_count, 0, vol_item)
        
        temp_item = QTableWidgetItem("")
        self.input_table.setItem(row_count, 1, temp_item)
    
    def remove_data_point(self):
        """Remove selected data point"""
        current_row = self.input_table.currentRow()
        if current_row >= 0:
            try:
                vol_pct = float(self.input_table.item(current_row, 0).text())
                if vol_pct in self.input_data:
                    del self.input_data[vol_pct]
            except (ValueError, AttributeError):
                pass
            self.input_table.removeRow(current_row)
    
    def clear_all_data(self):
        """Clear all input data"""
        self.input_data.clear()
        self.update_input_table()
    
    def convert_volume_to_weight_percent(self, vol_percents, densities):
        """
        Convert volume % to weight % using per-cut density values.
        
        If a single density is provided, it's used for all cuts.
        If per-cut densities are provided, they're used individually.
        
        Parameters:
        vol_percents: list of volume percentages [0, 10, 30, 50, 70, 90, 95]
        densities: float (constant) or list of floats (per-cut)
        
        Returns:
        list of weight percentages corresponding to input volume percentages
        """
        if not vol_percents:
            return []
        
        # Handle single density value (applies to all cuts)
        if isinstance(densities, (int, float)):
            density_values = [densities] * len(vol_percents)
        else:
            # Assume it's a list/dict of per-cut densities
            density_values = densities if isinstance(densities, list) else [densities] * len(vol_percents)
        
        # Cumulative volume and weight calculations
        # Assume linear density variation between cuts
        weight_percents = []
        cumulative_weight = 0.0
        total_mass = 0.0
        
        # First pass: calculate total mass based on volumes and densities
        for i in range(len(vol_percents)):
            if i == 0:
                vol_slice = vol_percents[i]
            else:
                vol_slice = vol_percents[i] - vol_percents[i-1]
            
            density = density_values[i] if i < len(density_values) else density_values[-1]
            mass_slice = vol_slice * density  # mass is proportional to volume * density
            total_mass += mass_slice
        
        # Second pass: convert to weight percentages
        cumulative_mass = 0.0
        for i in range(len(vol_percents)):
            if i == 0:
                vol_slice = vol_percents[i]
            else:
                vol_slice = vol_percents[i] - vol_percents[i-1]
            
            density = density_values[i] if i < len(density_values) else density_values[-1]
            mass_slice = vol_slice * density
            cumulative_mass += mass_slice
            
            if total_mass > 0:
                wt_pct = (cumulative_mass / total_mass) * 100
            else:
                wt_pct = vol_percents[i]  # Fallback to vol% if calculation fails
            
            weight_percents.append(wt_pct)
        
        return weight_percents
    
    def convert_weight_to_volume_percent(self, wt_percents, densities):
        """
        Convert weight % to volume % using per-cut density values.
        
        Parameters:
        wt_percents: list of weight percentages
        densities: float (constant) or list of floats (per-cut)
        
        Returns:
        list of volume percentages corresponding to input weight percentages
        """
        if not wt_percents:
            return []
        
        # Handle single density value
        if isinstance(densities, (int, float)):
            density_values = [densities] * len(wt_percents)
        else:
            density_values = densities if isinstance(densities, list) else [densities] * len(wt_percents)
        
        # Reverse calculation: from weight% to volume%
        volume_percents = []
        cumulative_volume = 0.0
        total_volume = 0.0
        
        # First pass: calculate total volume
        for i in range(len(wt_percents)):
            if i == 0:
                wt_slice = wt_percents[i]
            else:
                wt_slice = wt_percents[i] - wt_percents[i-1]
            
            density = density_values[i] if i < len(density_values) else density_values[-1]
            # mass ‚àù wt_slice, volume = mass/density
            vol_slice = wt_slice / density if density > 0 else wt_slice
            total_volume += vol_slice
        
        # Second pass: convert to volume percentages
        cumulative_volume = 0.0
        for i in range(len(wt_percents)):
            if i == 0:
                wt_slice = wt_percents[i]
            else:
                wt_slice = wt_percents[i] - wt_percents[i-1]
            
            density = density_values[i] if i < len(density_values) else density_values[-1]
            vol_slice = wt_slice / density if density > 0 else wt_slice
            cumulative_volume += vol_slice
            
            if total_volume > 0:
                vol_pct = (cumulative_volume / total_volume) * 100
            else:
                vol_pct = wt_percents[i]  # Fallback
            
            volume_percents.append(vol_pct)
        
        return volume_percents
    
    def calculate_conversions(self):
        """Perform the distillation curve conversions"""
        # Collect all input data
        self.input_data.clear()
        self.input_densities = {}  # Store per-cut density values if available
        
        for row in range(self.input_table.rowCount()):
            try:
                vol_item = self.input_table.item(row, 0)
                temp_item = self.input_table.item(row, 1)
                
                if vol_item and temp_item and temp_item.text():
                    vol_pct = float(vol_item.text())
                    temp = float(temp_item.text())
                    self.input_data[vol_pct] = temp
            except (ValueError, AttributeError):
                continue
        
        # Validate input
        if len(self.input_data) < 3:
            QMessageBox.warning(self, "Insufficient Data", 
                              "Please enter at least 3 data points for conversion.")
            return
        
        # Get input type and basis
        input_type = self.input_type_combo.currentText()
        basis = self.basis_combo.currentText()
        
        # Extract the actual type (D86, D2887, or TBP) from the combo box text
        if "D86" in input_type:
            input_type_str = "D86"
        elif "D2887" in input_type or "SimDist" in input_type:
            input_type_str = "D2887"
        elif "TBP" in input_type:
            input_type_str = "TBP"
        else:
            input_type_str = "D86"  # Default to D86
        
        # Prepare data for Oil class (needs list of [vol%, temp] pairs)
        input_data_list = [[vol, temp] for vol, temp in sorted(self.input_data.items())]
        
        # Handle basis conversion if needed
        # If input is in Weight %, convert to Volume % for internal calculations
        if "Weight" in basis:
            vol_percents = [point[0] for point in input_data_list]
            # Convert weight% to volume% using available density values
            # Try to use per-cut densities if available, otherwise use average density
            if self.input_densities:
                density_list = [self.input_densities.get(vp, self.density) for vp in vol_percents]
            else:
                density_list = [self.density] * len(vol_percents)
            
            vol_percents_converted = self.convert_weight_to_volume_percent(vol_percents, density_list)
            
            # Update input_data_list with converted volume percentages
            input_data_list = [[vol_percents_converted[i], input_data_list[i][1]] 
                              for i in range(len(input_data_list))]
        
        try:
            # Create Oil object with the specified input type
            self.oil_object = Oil(input_data_list, Density=self.density, input_type=input_type_str)
            
            # Store the input type for proper labeling
            self.current_input_type = input_type
            
            # Update displays
            self.update_plot()
            self.update_results_table()
            self.update_properties_table()
            
            # Enable export buttons
            self.export_csv_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)
            self.export_json_btn.setEnabled(True)
            
            QMessageBox.information(self, "Success", 
                                  "Conversions calculated successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Calculation Error", 
                               f"Error during conversion:\n{str(e)}")
    
    def update_plot(self):
        """Update the plot with conversion results"""
        if not self.oil_object:
            return
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # Generate smooth curves
        vol_range = np.linspace(0, 100, 200)
        
        input_type = self.input_type_combo.currentText()
        
        # Determine which interpolator represents the actual input data
        # Since Oil class treats input as D86, we need to identify what was actually input
        if "D86" in input_type:
            input_interp = self.oil_object.D86_interp
            input_label = 'üîµ'
        elif "D2887" in input_type or "SimDis" in input_type:
            input_interp = self.oil_object.D2887_interp
            input_label = 'üî¥'
        else:  # TBP
            input_interp = self.oil_object.TBP_interp
            input_label = 'üü¢'
        
        # Plot input data points
        input_vols = sorted(self.input_data.keys())
        input_temps = [self.input_data[v] for v in input_vols]
        ax.scatter(input_vols, input_temps, s=100, c='black', marker='o', 
                  label=f'Input: {input_type.split()[0]}', zorder=5, edgecolors='white', linewidths=1.5)
        
        # Plot conversions with enhanced styling
        if self.convert_d86_check.isChecked():
            d86_temps = [self.oil_object.D86_interp(v) for v in vol_range]
            ax.plot(vol_range, d86_temps, 'b-', linewidth=2.5, label='D86', alpha=0.9)
        
        if self.convert_d2887_check.isChecked():
            d2887_temps = [self.oil_object.D2887_interp(v) for v in vol_range]
            ax.plot(vol_range, d2887_temps, 'r--', linewidth=2.5, label='D2887 (SimDis)', alpha=0.9)
        
        if self.convert_tbp_check.isChecked():
            tbp_temps = [self.oil_object.TBP_interp(v) for v in vol_range]
            ax.plot(vol_range, tbp_temps, 'g-.', linewidth=2.5, label='TBP (API)', alpha=0.9)
        
        # Professional styling
        ax.set_xlabel('Volume % Distilled', fontsize=13, fontweight='bold')
        ax.set_ylabel('Temperature (¬∞C)', fontsize=13, fontweight='bold')
        ax.set_title('Distillation Curve Interconversion', fontsize=15, fontweight='bold', pad=15)
        
        # Set x-axis limits and ticks
        ax.set_xlim(0, 100)
        ax.set_xticks(np.arange(0, 101, 10))  # Major ticks every 10%
        ax.set_xticks(np.arange(0, 101, 5), minor=True)  # Minor ticks every 5%
        
        # Set y-axis to auto-scale with nice intervals
        ax.yaxis.set_major_locator(MaxNLocator(10))  # ~10 major divisions
        ax.yaxis.set_minor_locator(AutoMinorLocator(2))  # 2 minor ticks between majors
        
        # Enhanced grid styling
        # Major vertical grid lines at key distillation points (10, 30, 50, 70, 90%)
        for x_val in [10, 30, 50, 70, 90]:
            ax.axvline(x=x_val, color='#4682B4', linestyle='--', linewidth=1.5, alpha=0.8, zorder=0)
        
        # Minor vertical grid line at 95%
        ax.axvline(x=95, color='#4682B4', linestyle='--', linewidth=1.0, alpha=0.6, zorder=0)
        
        # Minor horizontal grid lines (temperature)
        ax.grid(True, which='minor', axis='y', color='#4682B4', linestyle='--', 
                linewidth=0.7, alpha=0.5, zorder=0)
        
        # Major grid lines (lighter, dotted for subtlety)
        ax.grid(True, which='major', axis='both', color='#708090', linestyle=':', 
                linewidth=0.8, alpha=0.5, zorder=0)
        
        # Legend with better styling
        legend = ax.legend(loc='lower right', fontsize=11, frameon=True, 
                          shadow=True, fancybox=True, framealpha=0.98)
        legend.get_frame().set_facecolor('#FFFEF7')
        legend.get_frame().set_edgecolor('#8B7355')
        
        # Add paper-like background color (soft cream/off-white)
        ax.set_facecolor('#FFFEF7')
        self.figure.patch.set_facecolor('#FFF8E7')
        
        # Enhance tick labels
        ax.tick_params(axis='both', which='major', labelsize=10, length=6, width=1.2)
        ax.tick_params(axis='both', which='minor', length=3, width=0.8)
        
        # Add border styling
        for spine in ax.spines.values():
            spine.set_linewidth(1.2)
            spine.set_color('#8B7355')
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def update_results_table(self):
        """Update the results table with conversion data"""
        if not self.oil_object:
            return
        
        # Use standard points for display
        display_points = [0, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95, 100]
        
        self.results_table.setRowCount(len(display_points))
        
        input_type = self.input_type_combo.currentText()
        
        for i, vol in enumerate(display_points):
            # Volume %
            self.results_table.setItem(i, 0, QTableWidgetItem(f"{vol:.1f}"))
            
            # Input temperature (if exists)
            if vol in self.input_data:
                input_temp = f"{self.input_data[vol]:.2f}"
            else:
                input_temp = "-"
            self.results_table.setItem(i, 1, QTableWidgetItem(input_temp))
            
            # D86
            try:
                d86_temp = self.oil_object.D86_interp(vol)
                self.results_table.setItem(i, 2, QTableWidgetItem(f"{d86_temp:.2f}"))
            except:
                self.results_table.setItem(i, 2, QTableWidgetItem("-"))
            
            # D2887
            try:
                d2887_temp = self.oil_object.D2887_interp(vol)
                self.results_table.setItem(i, 3, QTableWidgetItem(f"{d2887_temp:.2f}"))
            except:
                self.results_table.setItem(i, 3, QTableWidgetItem("-"))
    
    def update_properties_table(self):
        """Update the properties table with calculated properties"""
        if not self.oil_object:
            return
        
        properties = [
            ("Density", f"{self.density:.1f} kg/m¬≥"),
            ("Specific Gravity", f"{self.oil_object.SG:.4f}"),
            ("VABP (Volume Average BP)", f"{self.oil_object.VABP:.2f} ¬∞F"),
            ("MeABP (Mean Average BP)", f"{self.oil_object.MeABP:.2f} ¬∞F"),
            ("Watson K Factor", f"{self.oil_object.WatsonK:.2f}"),
        ]
        
        self.properties_table.setRowCount(len(properties))
        
        for i, (prop_name, prop_value) in enumerate(properties):
            self.properties_table.setItem(i, 0, QTableWidgetItem(prop_name))
            self.properties_table.setItem(i, 1, QTableWidgetItem(prop_value))
    
    def export_csv(self):
        """Export results to CSV file"""
        if not self.oil_object:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Save CSV", "", "CSV Files (*.csv)")
        if not file_path:
            return
        
        try:
            import csv
            
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                
                # Write header
                writer.writerow(["Distillation Curve Interconversion Results"])
                writer.writerow([])
                writer.writerow(["Density (kg/m¬≥):", self.density])
                writer.writerow(["VABP (¬∞F):", f"{self.oil_object.VABP:.2f}"])
                writer.writerow(["MeABP (¬∞F):", f"{self.oil_object.MeABP:.2f}"])
                writer.writerow(["Watson K:", f"{self.oil_object.WatsonK:.2f}"])
                writer.writerow([])
                
                # Write data table
                writer.writerow(["Vol %", "D86 (¬∞C)", "D2887 (¬∞C)", "TBP API (¬∞C)"])
                
                display_points = list(range(0, 101, 5))
                for vol in display_points:
                    row = [
                        vol,
                        f"{float(self.oil_object.D86_interp(vol)):.2f}",
                        f"{float(self.oil_object.D2887_interp(vol)):.2f}",
                        f"{float(self.oil_object.TBP_interp(vol)):.2f}"
                    ]
                    writer.writerow(row)
            
            QMessageBox.information(self, "Export Successful", 
                                  f"Data exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", 
                               f"Error exporting data:\n{str(e)}")
    
    def export_excel(self):
        """Export results to Excel file"""
        if not self.oil_object:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Distillation Curves"
            
            # Title
            ws['A1'] = "Distillation Curve Interconversion Results"
            ws['A1'].font = Font(size=14, bold=True)
            
            # Properties
            row = 3
            ws[f'A{row}'] = "Density (kg/m¬≥):"
            ws[f'B{row}'] = self.density
            row += 1
            ws[f'A{row}'] = "VABP (¬∞F):"
            ws[f'B{row}'] = round(self.oil_object.VABP, 2)
            row += 1
            ws[f'A{row}'] = "MeABP (¬∞F):"
            ws[f'B{row}'] = round(self.oil_object.MeABP, 2)
            row += 1
            ws[f'A{row}'] = "Watson K:"
            ws[f'B{row}'] = round(self.oil_object.WatsonK, 2)
            
            # Data table
            row = 9
            headers = ["Vol %", "D86 (¬∞C)", "D2887 (¬∞C)", "TBP API (¬∞C)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            display_points = list(range(0, 101, 5))
            for vol in display_points:
                row += 1
                ws.cell(row=row, column=1, value=vol)
                ws.cell(row=row, column=2, value=round(float(self.oil_object.D86_interp(vol)), 2))
                ws.cell(row=row, column=3, value=round(float(self.oil_object.D2887_interp(vol)), 2))
                ws.cell(row=row, column=4, value=round(float(self.oil_object.TBP_interp(vol)), 2))
            
            # Adjust column widths
            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 18
            
            wb.save(file_path)
            
            QMessageBox.information(self, "Export Successful", 
                                  f"Data exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", 
                               f"Error exporting data:\n{str(e)}")
    
    def export_json(self):
        """Export results to JSON file"""
        if not self.oil_object:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Save JSON", "", "JSON Files (*.json)")
        if not file_path:
            return
        
        try:
            import json
            
            # Build data structure
            display_points = list(range(0, 101, 5))
            
            data = {
                "metadata": {
                    "title": "Distillation Curve Interconversion Results",
                    "export_date": "2025-10-09",
                    "software": "Distillation Curve Interconversion Tool v1.0"
                },
                "material_properties": {
                    "density_kg_m3": self.density,
                    "VABP_F": round(self.oil_object.VABP, 2),
                    "MeABP_F": round(self.oil_object.MeABP, 2),
                    "Watson_K": round(self.oil_object.WatsonK, 2)
                },
                "distillation_curves": {
                    "D86": {},
                    "D2887": {},
                    "TBP_API": {}
                },
                "conversion_methods": {
                    "D86_to_TBP_API": "API Technical Data Book (1997) power-law correlations",
                    "D86_to_D2887": "Inverse of API Procedure 3A3.2",
                    "interpolation": "PCHIP (Piecewise Cubic Hermite Interpolating Polynomial)"
                }
            }
            
            # Add distillation data
            for vol in display_points:
                vol_str = str(vol)
                data["distillation_curves"]["D86"][vol_str] = {
                    "volume_percent": vol,
                    "temperature_C": round(float(self.oil_object.D86_interp(vol)), 2)
                }
                data["distillation_curves"]["D2887"][vol_str] = {
                    "volume_percent": vol,
                    "temperature_C": round(float(self.oil_object.D2887_interp(vol)), 2)
                }
                data["distillation_curves"]["TBP_API"][vol_str] = {
                    "volume_percent": vol,
                    "temperature_C": round(float(self.oil_object.TBP_interp(vol)), 2)
                }
            
            # Write to file with pretty formatting
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            QMessageBox.information(self, "Export Successful", 
                                  f"Data exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", 
                               f"Error exporting data:\n{str(e)}")
    
    def import_csv(self):
        """Import distillation data from CSV file"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Import CSV File", "", "CSV Files (*.csv);;All Files (*.*)")
        if not file_path:
            return
        
        try:
            import pandas as pd
            
            # Read CSV file
            df = pd.read_csv(file_path)
            
            # Auto-detect columns (case-insensitive)
            columns_lower = [col.lower() for col in df.columns]
            
            # Find volume column
            vol_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['vol', 'percentage', '%', 'cut']):
                    vol_col = df.columns[i]
                    break
            
            # Find temperature column
            temp_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['temp', 'temperature', '¬∞c', 'celsius', 'c', 'deg']):
                    temp_col = df.columns[i]
                    break
            
            # Find density column (optional)
            dens_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['dens', 'density', 'kg/m', 'kg/m3']):
                    dens_col = df.columns[i]
                    break
            
            if vol_col is None or temp_col is None:
                available = ", ".join(df.columns)
                QMessageBox.warning(self, "Import Error",
                    f"Could not auto-detect required columns.\n\n"
                    f"Looking for: Volume (%) and Temperature (¬∞C) columns\n\n"
                    f"Available columns: {available}\n\n"
                    f"Please rename columns to include 'Vol' and 'Temp'")
                return
            
            # Clear existing data
            self.input_table.setRowCount(0)
            
            # Import data and store per-cut densities if available
            self.input_densities = {}  # Store per-cut density values
            for idx, row in df.iterrows():
                try:
                    vol = float(row[vol_col])
                    temp = float(row[temp_col])
                    
                    # Add row to table
                    row_pos = self.input_table.rowCount()
                    self.input_table.insertRow(row_pos)
                    
                    vol_item = QTableWidgetItem(str(vol))
                    temp_item = QTableWidgetItem(str(temp))
                    
                    self.input_table.setItem(row_pos, 0, vol_item)
                    self.input_table.setItem(row_pos, 1, temp_item)
                    
                    # Store per-cut density if available
                    if dens_col is not None:
                        try:
                            dens = float(row[dens_col])
                            if 600 <= dens <= 1200:
                                self.input_densities[vol] = dens
                        except (ValueError, TypeError):
                            pass
                except (ValueError, KeyError):
                    # Skip rows with invalid data
                    continue
            
            # Update density field with average of per-cut densities
            if self.input_densities:
                avg_density = sum(self.input_densities.values()) / len(self.input_densities)
                if 600 <= avg_density <= 1200:
                    self.density_input.setValue(avg_density)
            elif dens_col is not None and len(df) > 0:
                # Fallback: use first density if per-cut data not available
                try:
                    density = float(df[dens_col].iloc[0])
                    if 600 <= density <= 1200:
                        self.density_input.setValue(density)
                except (ValueError, TypeError):
                    pass
            
            # Auto-detect input type from filename
            filename_lower = file_path.lower()
            if 'd2887' in filename_lower:
                self.input_type_combo.setCurrentText("D2887")
            elif 'tbp' in filename_lower or 'atm' in filename_lower:
                self.input_type_combo.setCurrentText("TBP")
            else:
                self.input_type_combo.setCurrentText("D86")
            
            rows_imported = self.input_table.rowCount()
            density_info = f" ({len(self.input_densities)} cuts with per-cut density)" if self.input_densities else ""
            QMessageBox.information(self, "Import Successful",
                f"Imported {rows_imported} data points from CSV{density_info}")

            
        except Exception as e:
            QMessageBox.critical(self, "Import Error",
                f"Error importing CSV file:\n{str(e)}")
    
    def import_excel(self):
        """Import distillation data from Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Excel File", "", "Excel Files (*.xlsx *.xls);;All Files (*.*)")
        if not file_path:
            return
        
        try:
            import pandas as pd
            
            # Read Excel file (first sheet)
            df = pd.read_excel(file_path)
            
            # Auto-detect columns (case-insensitive)
            columns_lower = [col.lower() for col in df.columns]
            
            # Find volume column
            vol_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['vol', 'percentage', '%', 'cut']):
                    vol_col = df.columns[i]
                    break
            
            # Find temperature column
            temp_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['temp', 'temperature', '¬∞c', 'celsius', 'c', 'deg']):
                    temp_col = df.columns[i]
                    break
            
            # Find density column (optional)
            dens_col = None
            for i, col_lower in enumerate(columns_lower):
                if any(x in col_lower for x in ['dens', 'density', 'kg/m', 'kg/m3']):
                    dens_col = df.columns[i]
                    break
            
            if vol_col is None or temp_col is None:
                available = ", ".join(df.columns)
                QMessageBox.warning(self, "Import Error",
                    f"Could not auto-detect required columns.\n\n"
                    f"Looking for: Volume (%) and Temperature (¬∞C) columns\n\n"
                    f"Available columns: {available}\n\n"
                    f"Please rename columns to include 'Vol' and 'Temp'")
                return
            
            # Clear existing data
            self.input_table.setRowCount(0)
            
            # Import data and store per-cut densities if available
            self.input_densities = {}  # Store per-cut density values
            for idx, row in df.iterrows():
                try:
                    vol = float(row[vol_col])
                    temp = float(row[temp_col])
                    
                    # Add row to table
                    row_pos = self.input_table.rowCount()
                    self.input_table.insertRow(row_pos)
                    
                    vol_item = QTableWidgetItem(str(vol))
                    temp_item = QTableWidgetItem(str(temp))
                    
                    self.input_table.setItem(row_pos, 0, vol_item)
                    self.input_table.setItem(row_pos, 1, temp_item)
                    
                    # Store per-cut density if available
                    if dens_col is not None:
                        try:
                            dens = float(row[dens_col])
                            if 600 <= dens <= 1200:
                                self.input_densities[vol] = dens
                        except (ValueError, TypeError):
                            pass
                except (ValueError, KeyError):
                    # Skip rows with invalid data
                    continue
            
            # Update density field with average of per-cut densities
            if self.input_densities:
                avg_density = sum(self.input_densities.values()) / len(self.input_densities)
                if 600 <= avg_density <= 1200:
                    self.density_input.setValue(avg_density)
            elif dens_col is not None and len(df) > 0:
                # Fallback: use first density if per-cut data not available
                try:
                    density = float(df[dens_col].iloc[0])
                    if 600 <= density <= 1200:
                        self.density_input.setValue(density)
                except (ValueError, TypeError):
                    pass
            
            # Auto-detect input type from filename
            filename_lower = file_path.lower()
            if 'd2887' in filename_lower:
                self.input_type_combo.setCurrentText("D2887")
            elif 'tbp' in filename_lower or 'atm' in filename_lower:
                self.input_type_combo.setCurrentText("TBP")
            else:
                self.input_type_combo.setCurrentText("D86")
            
            rows_imported = self.input_table.rowCount()
            density_info = f" ({len(self.input_densities)} cuts with per-cut density)" if self.input_densities else ""
            QMessageBox.information(self, "Import Successful",
                f"Imported {rows_imported} data points from Excel{density_info}")
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error",
                f"Error importing Excel file:\n{str(e)}")


def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Use Windows 11/10 native style - automatically uses current Windows styling
    # Try different Windows styles in order of preference
    try:
        # PySide6 on Windows typically uses 'Windows' style by default
        # which automatically adapts to Windows 10/11
        app.setStyle('Windows')
    except:
        # Fallback to windowsvista if Windows style not available
        try:
            app.setStyle('windowsvista')
        except:
            pass  # Use default style
    
    window = DistillationConverterGUI()
    window.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
